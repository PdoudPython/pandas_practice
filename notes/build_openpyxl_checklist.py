import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Learning Checklist"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
STEP_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
NOTES_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

col_widths = {"A": 8, "B": 42, "C": 34, "D": 55}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

row = 1

# Title
ws.merge_cells(f"A{row}:D{row}")
c = ws.cell(row=row, column=1, value="openpyxl Learning Checklist")
c.font = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[row].height = 30
for col in range(1, 5):
    ws.cell(row=row, column=col).fill = HEADER_FILL
row += 1

ws.merge_cells(f"A{row}:D{row}")
c = ws.cell(row=row, column=1,
            value="Work top to bottom, building one small script per stage. Check off each item, "
                  "write your own working syntax in the Notes row once you've tested it — not before.")
c.font = Font(name=FONT_NAME, size=10, italic=True, color="404040")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[row].height = 32
row += 2

headers = ["Done", "Checklist Item", "Key Objects / Methods", "Your Notes"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=i, value=h)
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
ws.row_dimensions[row].height = 20
header_row = row
row += 1

dv = DataValidation(type="list", formula1='"☐,☑"', allow_blank=True)
ws.add_data_validation(dv)

steps = [
    ("STEP 1 — WORKBOOKS & WORKSHEETS", [
        ("Create a new workbook and grab the active sheet", "openpyxl.Workbook(), wb.active"),
        ("Rename a sheet and create additional sheets", "ws.title, wb.create_sheet()"),
        ("Switch between / list existing sheets", "wb.sheetnames, wb['SheetName']"),
        ("Delete a sheet", "wb.remove(ws)"),
        ("Open an existing .xlsx file", "openpyxl.load_workbook()"),
        ("Save a workbook to disk", "wb.save('file.xlsx')"),
    ]),
    ("STEP 2 — READING & WRITING CELLS", [
        ("Write a value to a specific cell (two addressing styles)", "ws['A1'] = value, ws.cell(row=, column=, value=)"),
        ("Read a value back out of a cell", "ws['A1'].value"),
        ("Loop over a range of cells", "ws.iter_rows(), ws.iter_cols()"),
        ("Append a full row at once", "ws.append([...])"),
        ("Understand formulas vs. cached values", "data_only=True on load_workbook()"),
    ]),
    ("STEP 3 — STYLING CELLS", [
        ("Set font (name, size, bold, italic, color)", "openpyxl.styles.Font"),
        ("Set fill/background color", "openpyxl.styles.PatternFill"),
        ("Set text alignment and wrapping", "openpyxl.styles.Alignment"),
        ("Add cell borders", "openpyxl.styles.Border, Side"),
        ("Apply a number format (currency, %, dates)", "cell.number_format"),
    ]),
    ("STEP 4 — ROWS, COLUMNS & LAYOUT", [
        ("Set column width and row height", "ws.column_dimensions, ws.row_dimensions"),
        ("Merge and unmerge cells", "ws.merge_cells(), ws.unmerge_cells()"),
        ("Freeze header rows/columns", "ws.freeze_panes"),
        ("Insert or delete rows/columns", "ws.insert_rows(), ws.delete_cols()"),
        ("Hide gridlines or hide rows/columns", "ws.sheet_view.showGridLines, dimension.hidden"),
    ]),
    ("STEP 5 — FORMULAS", [
        ("Write a formula into a cell as a string", "ws['B2'] = '=SUM(B3:B10)'"),
        ("Understand openpyxl never calculates formulas itself", "requires Excel/LibreOffice to evaluate"),
        ("Know which functions need an _xlfn. prefix", "e.g. _xlfn.TEXTJOIN, _xlfn.IFS"),
        ("Recalculate a file headlessly (e.g. via LibreOffice)", "soffice --headless --convert-to xlsx"),
    ]),
    ("STEP 6 — DATA VALIDATION & INTERACTIVITY", [
        ("Add a dropdown list to a cell/range", "openpyxl.worksheet.datavalidation.DataValidation"),
        ("Restrict input to a number range or date", "DataValidation(type='whole'/'date', ...)"),
        ("Add conditional formatting", "openpyxl.formatting.rule"),
    ]),
    ("STEP 7 — CHARTS & VISUALS", [
        ("Build a bar/line/pie chart from cell data", "openpyxl.chart.BarChart/LineChart/PieChart"),
        ("Define chart data and category ranges", "openpyxl.chart.Reference"),
        ("Insert the chart onto a worksheet", "ws.add_chart(chart, 'E2')"),
        ("Insert an image into a sheet", "openpyxl.drawing.image.Image"),
    ]),
    ("STEP 8 — WORKING WITH REAL DATA", [
        ("Load a large dataset in read-only mode for speed", "load_workbook(read_only=True)"),
        ("Write large datasets efficiently", "ws.append() in write_only mode"),
        ("Convert between pandas DataFrames and worksheets", "pandas.read_excel(), df.to_excel()"),
        ("Handle multiple sheets/files in a loop", "wb.sheetnames + os.listdir()"),
    ]),
]

for step_title, items in steps:
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws.cell(row=row, column=1, value=step_title)
    cell.font = Font(name=FONT_NAME, size=12, bold=True, color="1F3864")
    cell.fill = STEP_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = STEP_FILL
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = 22
    row += 1

    for item_text, funcs in items:
        chk = ws.cell(row=row, column=1, value="☐")
        chk.font = Font(name=FONT_NAME, size=12)
        chk.alignment = Alignment(horizontal="center", vertical="center")
        chk.border = BORDER
        dv.add(chk)

        item_cell = ws.cell(row=row, column=2, value=item_text)
        item_cell.font = Font(name=FONT_NAME, size=10.5)
        item_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        item_cell.border = BORDER

        func_cell = ws.cell(row=row, column=3, value=funcs)
        func_cell.font = Font(name=FONT_NAME, size=10, italic=True, color="375623")
        func_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        func_cell.border = BORDER

        ws.row_dimensions[row].height = 26
        row += 1

        ws.merge_cells(f"B{row}:D{row}")
        notes_label = ws.cell(row=row, column=1, value="Notes")
        notes_label.font = Font(name=FONT_NAME, size=8, italic=True, color="8C8C8C")
        notes_label.alignment = Alignment(horizontal="center", vertical="top")
        notes_label.fill = NOTES_FILL
        notes_label.border = BORDER

        notes_cell = ws.cell(row=row, column=2, value="")
        notes_cell.fill = NOTES_FILL
        notes_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        notes_cell.border = BORDER
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = NOTES_FILL
            ws.cell(row=row, column=col).border = BORDER

        ws.row_dimensions[row].height = 45
        row += 1

ws.freeze_panes = f"A{header_row + 1}"

row += 1
ws.merge_cells(f"A{row}:D{row}")
c = ws.cell(row=row, column=1, value="GENERAL NOTES / PATTERNS I KEEP REUSING")
c.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
c.fill = HEADER_FILL
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
for col in range(1, 5):
    ws.cell(row=row, column=col).fill = HEADER_FILL
ws.row_dimensions[row].height = 22
row += 1

for _ in range(10):
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws.cell(row=row, column=1, value="")
    cell.fill = NOTES_FILL
    cell.border = BORDER
    ws.row_dimensions[row].height = 22
    row += 1

ws.sheet_view.showGridLines = False

wb.save("/mnt/user-data/outputs/openpyxl_learning_checklist.xlsx")
print("saved")
