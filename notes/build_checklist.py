import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cleaning Checklist"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
STEP_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
NOTES_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
DONE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Column layout
# A: Done (checkbox-style dropdown)
# B: Item
# C: Pandas function(s) / tool
# D: Your Notes (write-in space)
col_widths = {"A": 8, "B": 40, "C": 34, "D": 55}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

row = 1

# Title
ws.merge_cells(f"A{row}:D{row}")
c = ws.cell(row=row, column=1, value="Pandas Data-Cleaning Checklist")
c.font = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[row].height = 30
for col in range(1, 5):
    ws.cell(row=row, column=col).fill = HEADER_FILL
row += 1

ws.merge_cells(f"A{row}:D{row}")
c = ws.cell(row=row, column=1,
            value="Work top to bottom. Check off each item, jot the syntax/quirks you find in the Notes column, "
                  "then move on. Use the yellow rows to write freely — messy notes, gotchas, links, anything.")
c.font = Font(name=FONT_NAME, size=10, italic=True, color="404040")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[row].height = 32
row += 2

# Header row
headers = ["Done", "Checklist Item", "Pandas Function(s)", "Your Notes"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=i, value=h)
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
ws.row_dimensions[row].height = 20
header_row = row
row += 1

# Data validation dropdown for Done column
dv = DataValidation(type="list", formula1='"☐,☑"', allow_blank=True)
ws.add_data_validation(dv)

steps = [
    ("STEP 1 — LOAD & INSPECT", [
        ("Read the file in and confirm it loaded correctly", "pd.read_csv()"),
        ("Check shape (rows, columns)", "df.shape"),
        ("Preview first and last rows", "df.head(), df.tail()"),
        ("Check data types of every column", "df.dtypes / df.info()"),
        ("Get summary statistics for numeric columns", "df.describe()"),
        ("Check column names for consistency", "df.columns"),
    ]),
    ("STEP 2 — STRUCTURAL ISSUES", [
        ("Fix incorrect data types (e.g., dates read as strings)", "pd.to_datetime(), astype()"),
        ("Standardize column names (case, spacing, symbols)", "df.rename(), str.lower(), str.replace()"),
        ("Drop unnecessary columns", "df.drop(columns=[...])"),
        ("Reset or set a meaningful index", "df.reset_index(), df.set_index()"),
    ]),
    ("STEP 3 — MISSING DATA", [
        ("Find missing values and count them per column", "df.isna().sum()"),
        ("Decide: drop rows/columns with missing data", "df.dropna()"),
        ("Decide: fill missing values (mean/median/mode/constant)", "df.fillna()"),
        ("Forward-fill or back-fill for time series gaps", "df.ffill(), df.bfill()"),
    ]),
    ("STEP 4 — DUPLICATES", [
        ("Find duplicate rows", "df.duplicated()"),
        ("Remove duplicate rows", "df.drop_duplicates()"),
        ("Check for duplicate values in a key column (e.g., ID)", "df['col'].duplicated()"),
    ]),
    ("STEP 5 — INCONSISTENT VALUES", [
        ("Strip extra whitespace from text columns", "df['col'].str.strip()"),
        ("Standardize text case (upper/lower/title)", "str.lower(), str.upper(), str.title()"),
        ("Fix inconsistent category labels/typos", "df['col'].replace({...}), df['col'].map({...})"),
        ("Standardize units (currency, measurement, etc.)", "custom function + df.apply()"),
    ]),
    ("STEP 6 — OUTLIERS & INVALID VALUES", [
        ("Check numeric ranges for impossible values", "df['col'].min(), df['col'].max()"),
        ("Flag or filter outliers (e.g., z-score, IQR)", "df[(df['col'] > lower) & (df['col'] < upper)]"),
        ("Validate dates are within a sensible range", "boolean filtering on datetime column"),
    ]),
    ("STEP 7 — RESHAPING & COMBINING", [
        ("Reshape wide data to long (or vice versa)", "pd.melt(), df.pivot()"),
        ("Merge/join with another dataset", "pd.merge(), df.join()"),
        ("Group and aggregate data", "df.groupby().agg()"),
        ("Concatenate multiple files/dataframes", "pd.concat()"),
    ]),
    ("STEP 8 — FINAL VALIDATION", [
        ("Re-check shape, dtypes, and nulls after cleaning", "df.shape, df.info(), df.isna().sum()"),
        ("Spot-check a sample of rows manually", "df.sample(10)"),
        ("Save the cleaned dataset", "df.to_csv(), df.to_parquet()"),
    ]),
]

for step_title, items in steps:
    # Step header row
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws.cell(row=row, column=1, value=step_title)
    cell.font = Font(name=FONT_NAME, size=12, bold=True, color="1F3864")
    cell.fill = STEP_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = STEP_FILL
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = 22
    row += 1

    for item_text, funcs in items:
        # checkbox cell
        chk = ws.cell(row=row, column=1, value="☐")
        chk.font = Font(name=FONT_NAME, size=12)
        chk.alignment = Alignment(horizontal="center", vertical="center")
        chk.border = BORDER
        dv.add(chk)

        item_cell = ws.cell(row=row, column=2, value=item_text)
        item_cell.font = Font(name=FONT_NAME, size=10.5)
        item_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        item_cell.border = BORDER

        func_cell = ws.cell(row=row, column=3, value=funcs)
        func_cell.font = Font(name=FONT_NAME, size=10, italic=True, color="375623")
        func_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        func_cell.border = BORDER

        ws.row_dimensions[row].height = 26
        row += 1

        # notes row directly beneath each item — merged, tall, write-in space
        ws.merge_cells(f"B{row}:D{row}")
        notes_label = ws.cell(row=row, column=1, value="Notes")
        notes_label.font = Font(name=FONT_NAME, size=8, italic=True, color="8C8C8C")
        notes_label.alignment = Alignment(horizontal="center", vertical="top")
        notes_label.fill = NOTES_FILL
        notes_label.border = BORDER

        notes_cell = ws.cell(row=row, column=2, value="")
        notes_cell.fill = NOTES_FILL
        notes_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        notes_cell.border = BORDER
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = NOTES_FILL
            ws.cell(row=row, column=col).border = BORDER

        ws.row_dimensions[row].height = 45
        row += 1

# Freeze header
ws.freeze_panes = f"A{header_row + 1}"

# A blank "extra notes / cheat sheet" section at the bottom
row += 1
ws.merge_cells(f"A{row}:D{row}")
c = ws.cell(row=row, column=1, value="GENERAL NOTES / PATTERNS I KEEP REUSING")
c.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
c.fill = HEADER_FILL
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
for col in range(1, 5):
    ws.cell(row=row, column=col).fill = HEADER_FILL
ws.row_dimensions[row].height = 22
row += 1

for _ in range(10):
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws.cell(row=row, column=1, value="")
    cell.fill = NOTES_FILL
    cell.border = BORDER
    ws.row_dimensions[row].height = 22
    row += 1

ws.sheet_view.showGridLines = False

wb.save("/mnt/user-data/outputs/pandas_cleaning_checklist.xlsx")
print("saved")
